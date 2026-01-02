"""
Excel File Analyzer - For understanding eval and attainment sheet structures
"""
import openpyxl
from openpyxl.utils import get_column_letter
import os

def analyze_excel_structure(file_path, max_rows=30, max_cols=20):
    """Analyze Excel file structure and return formatted info"""
    print(f"\n{'='*100}")
    print(f"FILE: {os.path.basename(file_path)}")
    print(f"{'='*100}")
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        print(f"Sheet Names: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
            
            # Print header info and data structure
            print(f"\nFirst {max_rows} rows:")
            for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(max_cols + 1, ws.max_column + 1)):
                    cell = ws.cell(row_idx, col_idx)
                    value = cell.value
                    
                    # Show formulas if present
                    if cell.data_type == 'f':
                        row_data.append(f"[FORMULA: {value}]")
                    else:
                        row_data.append(str(value) if value is not None else "")
                
                print(f"Row {row_idx:3d}: {' | '.join(row_data)}")
            
            print(f"\n... (Total {ws.max_row} rows)")
    
    except Exception as e:
        print(f"ERROR: {e}")

# Analyze all relevant files
print("\n" + "="*100)
print(" ANALYSIS OF EVAL SHEETS AND ATTAINMENT TEMPLATES ".center(100, "="))
print("="*100)

# 1. DEPT THEORY EVAL SHEETS
print("\n\n### DEPARTMENT THEORY EVAL SHEETS ###")
analyze_excel_structure("sample/input_R17/theory_eval/Dept_theory/C211_IA1_b1923_r17.xlsx", max_rows=20)
analyze_excel_structure("sample/input_R17/theory_eval/Dept_theory/C211_ia2_B2023_R17.xlsx", max_rows=20)
analyze_excel_structure("sample/input_R17/theory_eval/Dept_theory/C211_mod_B1923_R17.xlsx", max_rows=20)

# 2. S&H THEORY EVAL SHEETS
print("\n\n### S&H THEORY EVAL SHEETS ###")
analyze_excel_structure("sample/input_R17/theory_eval/S&H_theory/C101_ia1_B1923_R17.xlsx", max_rows=20)
analyze_excel_structure("sample/input_R17/theory_eval/S&H_theory/C101_ia2_B2023_R17.xlsx", max_rows=20)
analyze_excel_structure("sample/input_R17/theory_eval/S&H_theory/C101_mod_B1923_R17.xlsx", max_rows=20)

# 3. ANALYTICAL EVAL SHEETS
print("\n\n### ANALYTICAL EVAL SHEETS ###")
analyze_excel_structure("sample/input_R17/analytical_eval/S&H_analytical/C102_ia1_B1923_R17.xlsx", max_rows=20)
analyze_excel_structure("sample/input_R17/analytical_eval/S&H_analytical/C102_ia2_B2023_R17.xlsx", max_rows=20)
analyze_excel_structure("sample/input_R17/analytical_eval/S&H_analytical/C102_mod_B1923_R17.xlsx", max_rows=20)

# 4. LAB EVAL SHEET
print("\n\n### LAB EVAL SHEET ###")
analyze_excel_structure("sample/input_R17/lab_eval/C107_b19-23_r17.xlsx", max_rows=25)

# 5. PROJECT EVAL SHEETS
print("\n\n### PROJECT EVAL SHEETS ###")
analyze_excel_structure("sample/input_R17/proj_eval/C411_project_review1_b1923_r17.xlsx", max_rows=20)
analyze_excel_structure("sample/input_R17/proj_eval/C411_project_review2_b1923_r17.xlsx", max_rows=20)
analyze_excel_structure("sample/input_R17/proj_eval/C411_project_review3_b1923_r17.xlsx", max_rows=20)

# 6. ATTAINMENT TEMPLATES
print("\n\n" + "="*100)
print(" ATTAINMENT TEMPLATES ".center(100, "="))
print("="*100)

print("\n\n### DEPT THEORY TEMPLATE ###")
analyze_excel_structure("Attainment_Template/Reg_17/Dept THEORY template_ R17 V3 AtSheet.xlsx", max_rows=40, max_cols=30)

print("\n\n### DEPT THEORY ANALYTICAL TEMPLATE ###")
analyze_excel_structure("Attainment_Template/Reg_17/Dept THEORY Analytical template_R17 V3 AtSheet.xlsx", max_rows=40, max_cols=30)

print("\n\n### S&H THEORY TEMPLATE ###")
analyze_excel_structure("Attainment_Template/Reg_17/S&H THEORY template _R17 V3 AtSheet.xlsx", max_rows=40, max_cols=30)

print("\n\n### S&H ANALYTICAL TEMPLATE ###")
analyze_excel_structure("Attainment_Template/Reg_17/S&H THEORY template Analytical_R17 V3 AtSheet.xlsx", max_rows=40, max_cols=30)

print("\n\n### LAB TEMPLATE ###")
analyze_excel_structure("Attainment_Template/Reg_17/LAB template_R17 V3 AtSheet.xlsx", max_rows=40, max_cols=30)

print("\n\n### PROJECT TEMPLATE ###")
analyze_excel_structure("Attainment_Template/Reg_17/Project template_R17 V3 AtSheet.xlsx", max_rows=40, max_cols=30)

# 7. OUTPUT ATTAINMENT SHEETS (SAMPLES)
print("\n\n" + "="*100)
print(" SAMPLE OUTPUT ATTAINMENT SHEETS ".center(100, "="))
print("="*100)

print("\n\n### DEPT THEORY OUTPUT ###")
analyze_excel_structure("sample/output_R17/B19-23-C211-CS8491-COMPUTER ARCHITECTURE.xlsx", max_rows=40, max_cols=30)

print("\n\n### S&H THEORY OUTPUT ###")
analyze_excel_structure("sample/output_R17/C101 Communicative English R17 V3 AtSheet.xlsx", max_rows=40, max_cols=30)

print("\n\n### ANALYTICAL OUTPUT ###")
analyze_excel_structure("sample/output_R17/C102 ENGINEERING MATHEMATICS I-Analytical_R17 V3 AtSheet.xlsx", max_rows=40, max_cols=30)

print("\n\n### LAB OUTPUT ###")
analyze_excel_structure("sample/output_R17/C107 PROBLEM SOLVING AND PYTHON PROGRAMMING LABORATORY -R17 V3 AtSheet.xlsx", max_rows=40, max_cols=30)

print("\n\n### PROJECT OUTPUT ###")
analyze_excel_structure("sample/output_R17/C411_project_attainment_b1923_r17.xlsx", max_rows=40, max_cols=30)

print("\n\n" + "="*100)
print(" ANALYSIS COMPLETE ".center(100, "="))
print("="*100)
