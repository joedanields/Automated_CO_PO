"""
Data Parser Module
Extracts marks and validation fields from evaluation sheets
"""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union
from io import BytesIO
import re


class DataParser:
    """Parses evaluation sheets and extracts student marks data"""
    
    # Standard row positions for metadata (1-indexed)
    METADATA_ROWS = {
        'course_code': 2,
        'course_name': 3,
        'faculty_name': 4,
        'academic_year': 5,
        'class_info': 6,
        'regulation': 7,
        'total_students': 8,
        'assessment_name': 9
    }
    
    # Column for metadata values (C = 3)
    METADATA_COL = 3
    
    # Header rows
    QUESTION_ROW = 11  # Question/Assessment numbers
    CO_MAPPING_ROW = 12  # Which CO each question maps to
    MAX_MARKS_ROW = 13  # Maximum marks for each question/CO
    DATA_START_ROW = 14  # First student data row
    
    # Data columns
    SNO_COL = 1
    REG_NO_COL = 2
    NAME_COL = 3
    
    def __init__(self):
        """Initialize DataParser"""
        pass
    
    def load_workbook(self, file_source: Union[str, BytesIO]) -> openpyxl.Workbook:
        """
        Load Excel workbook from file path or file-like object
        
        Args:
            file_source: Path to Excel file or BytesIO object
            
        Returns:
            openpyxl Workbook object
        """
        if isinstance(file_source, BytesIO):
            file_source.seek(0)  # Reset file pointer to beginning
            return openpyxl.load_workbook(file_source, data_only=True)
        return openpyxl.load_workbook(file_source, data_only=True)
    
    def extract_validation_fields(self, file_source: Union[str, BytesIO]) -> Dict[str, str]:
        """
        Extract metadata/validation fields from evaluation sheet
        
        Args:
            file_source: Path to evaluation sheet or BytesIO object
            
        Returns:
            Dictionary with validation fields:
            {
                'course_code': 'C211',
                'course_name': 'COMPUTER ARCHITECTURE',
                'faculty_name': 'ANANTHI M',
                'academic_year': '2020-2021 (EVEN)',
                'class_info': 'B.TECH.IT (2ND YEAR)',
                'regulation': 'R2017 - AUC',
                'total_students': '62',
                'assessment_name': 'INTERNAL ASSESSMENT-1'
            }
        """
        wb = self.load_workbook(file_source)
        ws = wb.active
        
        fields = {}
        for field_name, row_num in self.METADATA_ROWS.items():
            cell_value = ws.cell(row=row_num, column=self.METADATA_COL).value
            fields[field_name] = str(cell_value).strip() if cell_value else ''
        
        wb.close()
        return fields
    
    def normalize_regulation(self, reg_string: str) -> str:
        """
        Normalize regulation string to standard format
        
        Args:
            reg_string: e.g., 'R2017 - AUC', 'R2021', 'Regulation 2017'
            
        Returns:
            Normalized format: 'R17', 'R21', 'R24'
        """
        reg_string = str(reg_string).upper()
        
        # Extract year from various formats
        match = re.search(r'R?20?(\d{2})', reg_string)
        if match:
            year = match.group(1)
            return f'R{year}'
        
        return reg_string
    
    def detect_assessment_type(self, assessment_name: str) -> str:
        """
        Detect assessment type from assessment name
        
        Args:
            assessment_name: e.g., 'INTERNAL ASSESSMENT-1', 'MODEL EXAM', 'LABORATORY'
            
        Returns:
            Type: 'IA1', 'IA2', 'Model', 'Lab', 'Review1', 'Review2', 'Review3', 'Integrated'
        """
        assessment_name = str(assessment_name).upper()
        
        if 'INTERNAL' in assessment_name or 'IA' in assessment_name:
            if '1' in assessment_name:
                return 'IA1'
            elif '2' in assessment_name:
                return 'IA2'
        elif 'MODEL' in assessment_name:
            return 'Model'
        elif 'LAB' in assessment_name or 'LABORATORY' in assessment_name:
            return 'Lab'
        elif 'PROJECT' in assessment_name or 'REVIEW' in assessment_name:
            if '1' in assessment_name:
                return 'Review1'
            elif '2' in assessment_name:
                return 'Review2'
            elif '3' in assessment_name:
                return 'Review3'
        elif 'INTEGRATED' in assessment_name:
            return 'Integrated'
        
        return 'Unknown'
    
    def find_co_columns(self, ws) -> List[Tuple[int, int]]:
        """
        Find columns containing CO totals (not individual question marks)
        
        Args:
            ws: Worksheet object
            
        Returns:
            List of tuples: [(column_index, co_number), ...]
        """
        co_columns = []
        
        # Check row 11 for 'CO' headers and row 12 for CO numbers
        for col in range(4, ws.max_column + 1):
            header = ws.cell(row=self.QUESTION_ROW, column=col).value
            co_num = ws.cell(row=self.CO_MAPPING_ROW, column=col).value
            
            if header and str(header).upper().strip() == 'CO':
                try:
                    co_number = int(float(str(co_num)))
                    co_columns.append((col, co_number))
                except (ValueError, TypeError):
                    pass
        
        return co_columns
    
    def extract_student_data(self, file_source: Union[str, BytesIO]) -> Dict[str, Dict]:
        """
        Extract student marks from evaluation sheet
        
        Args:
            file_source: Path to evaluation sheet or BytesIO object
            
        Returns:
            Dictionary with student data:
            {
                '711719205002': {
                    'name': 'ADITHYA R',
                    'reg_no': '711719205002',
                    'co_marks': {1: 26, 2: 12},
                    'total': 38
                },
                ...
            }
        """
        wb = self.load_workbook(file_source)
        ws = wb.active
        
        # Find CO columns
        co_columns = self.find_co_columns(ws)
        
        # Find TOTAL column
        total_col = None
        for col in range(4, ws.max_column + 1):
            header = ws.cell(row=self.QUESTION_ROW, column=col).value
            if header and 'TOTAL' in str(header).upper():
                total_col = col
                break
        
        # Extract student data
        students = {}
        for row in range(self.DATA_START_ROW, ws.max_row + 1):
            reg_no = ws.cell(row=row, column=self.REG_NO_COL).value
            name = ws.cell(row=row, column=self.NAME_COL).value
            
            # Skip empty rows
            if not reg_no or not name:
                continue
            
            reg_no = str(reg_no).strip()
            name = str(name).strip()
            
            # Extract CO marks
            co_marks = {}
            for col, co_num in co_columns:
                mark = ws.cell(row=row, column=col).value
                try:
                    co_marks[co_num] = float(mark) if mark is not None else 0
                except (ValueError, TypeError):
                    co_marks[co_num] = 0
            
            # Extract total
            total = 0
            if total_col:
                total_val = ws.cell(row=row, column=total_col).value
                try:
                    total = float(total_val) if total_val is not None else 0
                except (ValueError, TypeError):
                    total = 0
            
            students[reg_no] = {
                'name': name,
                'reg_no': reg_no,
                'co_marks': co_marks,
                'total': total
            }
        
        wb.close()
        return students
    
    def extract_max_marks(self, file_source: Union[str, BytesIO]) -> Dict[str, Any]:
        """
        Extract maximum marks for each CO from evaluation sheet
        
        Args:
            file_source: Path to evaluation sheet or BytesIO object
            
        Returns:
            Dictionary with max marks:
            {
                'co_max': {1: 30, 2: 20},
                'total_max': 50
            }
        """
        wb = self.load_workbook(file_source)
        ws = wb.active
        
        co_columns = self.find_co_columns(ws)
        
        # Find TOTAL column
        total_col = None
        for col in range(4, ws.max_column + 1):
            header = ws.cell(row=self.QUESTION_ROW, column=col).value
            if header and 'TOTAL' in str(header).upper():
                total_col = col
                break
        
        # Extract max marks from row 13
        co_max = {}
        for col, co_num in co_columns:
            max_mark = ws.cell(row=self.MAX_MARKS_ROW, column=col).value
            try:
                co_max[co_num] = float(max_mark) if max_mark is not None else 0
            except (ValueError, TypeError):
                co_max[co_num] = 0
        
        total_max = 0
        if total_col:
            total_val = ws.cell(row=self.MAX_MARKS_ROW, column=total_col).value
            try:
                total_max = float(total_val) if total_val is not None else 0
            except (ValueError, TypeError):
                total_max = 0
        
        wb.close()
        return {
            'co_max': co_max,
            'total_max': total_max
        }
    
    def merge_eval_data(self, eval_data_list: List[Dict]) -> Dict[str, Dict]:
        """
        Merge student data from multiple evaluation sheets
        
        Args:
            eval_data_list: List of student data dictionaries from different assessments
            
        Returns:
            Merged student data with CO marks from all assessments
        """
        merged = {}
        
        for eval_data in eval_data_list:
            for reg_no, student_info in eval_data.items():
                if reg_no not in merged:
                    merged[reg_no] = {
                        'name': student_info['name'],
                        'reg_no': reg_no,
                        'co_marks': {}
                    }
                
                # Merge CO marks
                for co_num, mark in student_info['co_marks'].items():
                    if co_num not in merged[reg_no]['co_marks']:
                        merged[reg_no]['co_marks'][co_num] = mark
                    else:
                        # If CO already exists, keep the existing value or take max
                        # (depends on requirement - here we keep first)
                        pass
        
        return merged
    
    def get_all_students_sorted(self, student_data: Dict) -> List[Dict]:
        """
        Get sorted list of all students
        
        Args:
            student_data: Dictionary of student data
            
        Returns:
            List of student dictionaries sorted by registration number
        """
        students = list(student_data.values())
        students.sort(key=lambda x: x['reg_no'])
        return students


# Test the parser
if __name__ == '__main__':
    parser = DataParser()
    
    # Test with sample file
    test_file = 'sample/input_R17/theory_eval/Dept_theory/C211_IA1_b1923_r17.xlsx'
    
    print("=== Validation Fields ===")
    fields = parser.extract_validation_fields(test_file)
    for key, value in fields.items():
        print(f"  {key}: {value}")
    
    print(f"\nNormalized Regulation: {parser.normalize_regulation(fields['regulation'])}")
    print(f"Assessment Type: {parser.detect_assessment_type(fields['assessment_name'])}")
    
    print("\n=== Max Marks ===")
    max_marks = parser.extract_max_marks(test_file)
    print(f"  CO Max: {max_marks['co_max']}")
    print(f"  Total Max: {max_marks['total_max']}")
    
    print("\n=== Student Data (first 5) ===")
    students = parser.extract_student_data(test_file)
    for i, (reg_no, data) in enumerate(list(students.items())[:5]):
        print(f"  {reg_no}: {data['name']} - CO: {data['co_marks']}, Total: {data['total']}")
