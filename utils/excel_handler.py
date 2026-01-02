"""
Excel Handler Module
Handles reading/writing Excel files with formula preservation
"""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from copy import copy
import shutil
from .template_mapper import TemplateMapper
from .data_parser import DataParser


class ExcelHandler:
    """Handles Excel operations for attainment sheet generation"""
    
    # Template column mappings for R17 Dept Theory
    # Format: {CO_number: {assessment_type: column_index}}
    R17_DEPT_THEORY_MAPPING = {
        'data_start_row': 7,
        'reg_no_col': 2,
        'name_col': 3,
        'co_columns': {
            1: {'IA1': 4, 'Model': 5},
            2: {'IA1': 8, 'Model': 9},
            3: {'IA2': 12, 'Model': 13},
            4: {'IA2': 16, 'Model': 17},
            5: {'Model': 20}
        }
    }
    
    # R17 S&H Theory mapping (may be different)
    R17_SH_THEORY_MAPPING = {
        'data_start_row': 7,
        'reg_no_col': 2,
        'name_col': 3,
        'co_columns': {
            1: {'IA1': 4, 'Model': 5},
            2: {'IA1': 8, 'Model': 9},
            3: {'IA2': 12, 'Model': 13},
            4: {'IA2': 16, 'Model': 17},
            5: {'Model': 20}
        }
    }
    
    # R17 Lab mapping
    R17_LAB_MAPPING = {
        'data_start_row': 7,
        'reg_no_col': 2,
        'name_col': 3,
        'co_columns': {
            1: {'Lab': 4},
            2: {'Lab': 5},
            3: {'Lab': 6},
            4: {'Lab': 7},
            5: {'Lab': 8}
        }
    }
    
    # R17 Project mapping
    R17_PROJECT_MAPPING = {
        'data_start_row': 7,
        'reg_no_col': 2,
        'name_col': 3,
        'co_columns': {
            1: {'Review1': 4, 'Review2': 8, 'Review3': 12},
            2: {'Review1': 5, 'Review2': 9, 'Review3': 13},
            3: {'Review1': 6, 'Review2': 10, 'Review3': 14},
            4: {'Review1': 7, 'Review2': 11, 'Review3': 15},
            5: {'Review1': None, 'Review2': None, 'Review3': None}
        }
    }
    
    def __init__(self, base_path: str = None):
        """
        Initialize ExcelHandler
        
        Args:
            base_path: Base path to project directory
        """
        if base_path is None:
            self.base_path = Path(__file__).parent.parent
        else:
            self.base_path = Path(base_path)
        
        self.template_mapper = TemplateMapper(self.base_path)
        self.parser = DataParser()
    
    def get_mapping(self, regulation: str, category: str, dept_type: str) -> Dict:
        """
        Get column mapping for given parameters
        
        Args:
            regulation: R17, R21, R24
            category: theory, analytical, lab, project
            dept_type: dept, s&h
            
        Returns:
            Mapping dictionary
        """
        regulation = regulation.upper()
        category = category.lower()
        dept_type = dept_type.lower()
        
        if regulation == 'R17':
            if category == 'theory':
                if dept_type == 's&h':
                    return self.R17_SH_THEORY_MAPPING
                return self.R17_DEPT_THEORY_MAPPING
            elif category == 'analytical':
                return self.R17_DEPT_THEORY_MAPPING  # Same structure
            elif category == 'lab':
                return self.R17_LAB_MAPPING
            elif category == 'project':
                return self.R17_PROJECT_MAPPING
        
        # Default to R17 Dept Theory mapping
        return self.R17_DEPT_THEORY_MAPPING
    
    def load_template(self, regulation: str, category: str, dept_type: str = 'dept') -> Tuple[openpyxl.Workbook, Path]:
        """
        Load the appropriate template
        
        Args:
            regulation: R17, R21, R24
            category: theory, analytical, lab, project
            dept_type: dept, s&h, default
            
        Returns:
            Tuple of (Workbook, template_path)
        """
        template_path = self.template_mapper.get_template_path(
            regulation, category, dept_type
        )
        
        # Load with data_only=False to preserve formulas
        wb = openpyxl.load_workbook(template_path, data_only=False)
        return wb, template_path
    
    def copy_template(self, template_path: Path, output_path: Path) -> openpyxl.Workbook:
        """
        Create a copy of template file
        
        Args:
            template_path: Path to template file
            output_path: Path for output file
            
        Returns:
            Workbook object for the copied file
        """
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Copy file
        shutil.copy(template_path, output_path)
        
        # Open the copy
        return openpyxl.load_workbook(output_path, data_only=False)
    
    def fill_student_data(
        self,
        workbook: openpyxl.Workbook,
        student_data: Dict[str, Dict],
        eval_data: Dict[str, Dict[str, Dict]],
        mapping: Dict
    ) -> None:
        """
        Fill student data into template
        
        Args:
            workbook: Template workbook to fill
            student_data: Merged student data {reg_no: {name, reg_no, ...}}
            eval_data: Data from each eval {assessment_type: {reg_no: {co_marks}}}
            mapping: Column mapping for template
        """
        ws = workbook.active
        
        # Sort students by registration number
        sorted_students = sorted(student_data.items(), key=lambda x: x[0])
        
        data_start_row = mapping['data_start_row']
        reg_no_col = mapping['reg_no_col']
        name_col = mapping['name_col']
        co_columns = mapping['co_columns']
        
        for idx, (reg_no, student_info) in enumerate(sorted_students):
            row = data_start_row + idx
            
            # Fill reg no and name
            ws.cell(row=row, column=reg_no_col, value=reg_no)
            ws.cell(row=row, column=name_col, value=student_info['name'])
            
            # Fill CO marks from each assessment
            for co_num, assessment_cols in co_columns.items():
                for assessment_type, col in assessment_cols.items():
                    if col is None:
                        continue
                    
                    # Get marks from eval_data
                    if assessment_type in eval_data:
                        if reg_no in eval_data[assessment_type]:
                            co_marks = eval_data[assessment_type][reg_no].get('co_marks', {})
                            mark = co_marks.get(co_num, '')
                            if mark != '':
                                ws.cell(row=row, column=col, value=mark)
    
    def save_with_formulas(self, workbook: openpyxl.Workbook, output_path: Path) -> None:
        """
        Save workbook preserving formulas
        
        Args:
            workbook: Workbook to save
            output_path: Path to save to
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    
    def generate_attainment_sheet(
        self,
        regulation: str,
        category: str,
        dept_type: str,
        eval_files: Dict[str, str],
        output_path: str,
        course_info: Dict[str, str] = None
    ) -> Dict[str, Any]:
        """
        Main function to generate attainment sheet
        
        Args:
            regulation: R17, R21, R24
            category: theory, analytical, lab, project
            dept_type: dept, s&h, default
            eval_files: Dict mapping assessment type to file path
                        e.g., {'IA1': 'path/to/ia1.xlsx', 'IA2': 'path/to/ia2.xlsx', 'Model': 'path/to/model.xlsx'}
            output_path: Path for output file
            course_info: Optional course info to embed
            
        Returns:
            Result dictionary with success status and file path or error
        """
        try:
            # Load template
            template_wb, template_path = self.load_template(regulation, category, dept_type)
            
            # Get mapping
            mapping = self.get_mapping(regulation, category, dept_type)
            
            # Parse all evaluation files
            eval_data = {}
            merged_students = {}
            
            for assessment_type, file_path in eval_files.items():
                # Extract student data
                students = self.parser.extract_student_data(file_path)
                eval_data[assessment_type] = students
                
                # Merge student info
                for reg_no, student_info in students.items():
                    if reg_no not in merged_students:
                        merged_students[reg_no] = {
                            'name': student_info['name'],
                            'reg_no': reg_no
                        }
            
            # Copy template
            output_path = Path(output_path)
            workbook = self.copy_template(template_path, output_path)
            
            # Fill data
            self.fill_student_data(workbook, merged_students, eval_data, mapping)
            
            # Save
            self.save_with_formulas(workbook, output_path)
            
            return {
                'success': True,
                'file': str(output_path),
                'students_count': len(merged_students)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }


# Test the handler
if __name__ == '__main__':
    handler = ExcelHandler()
    
    # Test with sample files
    eval_files = {
        'IA1': 'sample/input_R17/theory_eval/Dept_theory/C211_IA1_b1923_r17.xlsx',
        'IA2': 'sample/input_R17/theory_eval/Dept_theory/C211_ia2_B2023_R17.xlsx',
        'Model': 'sample/input_R17/theory_eval/Dept_theory/C211_mod_B1923_R17.xlsx'
    }
    
    result = handler.generate_attainment_sheet(
        regulation='R17',
        category='theory',
        dept_type='dept',
        eval_files=eval_files,
        output_path='outputs/test_C211_attainment.xlsx'
    )
    
    print("Result:", result)
